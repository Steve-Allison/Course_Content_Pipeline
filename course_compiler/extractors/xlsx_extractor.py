"""
Microsoft Excel (.xlsx) file extractor implementation.
"""
from pathlib import Path
from typing import Dict, Any, List, Optional, Union, Tuple
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, PatternFill, Border, Alignment, Protection
import datetime
import json
import os

from .base import BaseExtractor, ExtractionError
from ..logging_utils import get_logger

logger = get_logger(__name__)

class XlsxExtractor(BaseExtractor):
    """Extracts content from Microsoft Excel (.xlsx) files."""
    
    def __init__(self, file_path: Path, config: Optional[Dict[str, Any]] = None):
        super().__init__(file_path, config)
        self.workbook: Optional[Workbook] = None
        self.include_formulas = self.config.get('include_formulas', True)
        self.include_formatting = self.config.get('include_formatting', True)
        self.max_rows_per_sheet = self.config.get('max_rows_per_sheet', 1000)
        self.max_columns_per_sheet = self.config.get('max_columns_per_sheet', 50)
    
    def load(self) -> None:
        """Load the Excel workbook."""
        try:
            # Load with data_only=False to access formulas if needed
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=not self.include_formulas,
                read_only=True
            )
        except Exception as e:
            raise ExtractionError(f"Failed to load Excel file: {e}") from e
    
    def extract(self) -> Dict[str, Any]:
        """
        Extract content from the Excel file.
        
        Returns:
            Dictionary containing worksheets, cells, and metadata
        """
        if self.workbook is None:
            self.load()
            
        try:
            metadata = self.get_metadata()
            
            content = {
                'metadata': metadata,
                'worksheets': [],
                'defined_names': self._extract_defined_names(),
                'file_type': 'xlsx',
                'sheet_count': len(self.workbook.sheetnames)
            }
            
            # Process each worksheet
            for sheet_name in self.workbook.sheetnames:
                try:
                    worksheet = self.workbook[sheet_name]
                    sheet_data = self._process_worksheet(worksheet)
                    content['worksheets'].append(sheet_data)
                except Exception as e:
                    logger.error(f"Error processing worksheet '{sheet_name}': {e}")
                    content['worksheets'].append({
                        'name': sheet_name,
                        'error': str(e)
                    })
            
            return content
            
        except Exception as e:
            error_msg = f"Error extracting content from {self.file_path}: {e}"
            logger.error(error_msg, exc_info=True)
            raise ExtractionError(error_msg) from e
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _process_worksheet(self, worksheet: Worksheet) -> Dict[str, Any]:
        """Process a single worksheet and extract its content."""
        try:
            # Get basic worksheet info
            sheet_data = {
                'name': worksheet.title,
                'dimensions': worksheet.dimensions,
                'active': worksheet == self.workbook.active,
                'sheet_state': worksheet.sheet_state,
                'views': [str(view) for view in worksheet.views],
                'merged_cells': [str(mc) for mc in worksheet.merged_cells.ranges],
                'data_validation': [str(dv) for dv in worksheet.data_validations.dataValidation],
                'auto_filter': str(worksheet.auto_filter) if worksheet.auto_filter else None,
                'print_options': {k: v for k, v in worksheet.print_options.items()},
                'page_setup': {k: v for k, v in worksheet.page_setup.items()},
                'header_footer': {
                    'header': str(worksheet.header_footer.header),
                    'footer': str(worksheet.header_footer.footer),
                    'align_with_margins': worksheet.header_footer.alignWithMargins,
                    'scale_with_doc': worksheet.header_footer.scaleWithDoc,
                },
                'rows': [],
                'tables': [],
                'charts': [],
                'images': [],
                'row_count': 0,
                'column_count': 0,
            }
            
            # Process cells
            rows_data = []
            max_col = 0
            
            # Limit the number of rows to process
            max_row = min(worksheet.max_row, self.max_rows_per_sheet)
            
            for row_idx, row in enumerate(worksheet.iter_rows(
                min_row=1, 
                max_row=max_row,
                max_col=self.max_columns_per_sheet,
                values_only=False
            ), 1):
                row_data = []
                for cell in row:
                    if cell.value is not None:
                        max_col = max(max_col, cell.column)
                        cell_data = self._process_cell(cell)
                        row_data.append(cell_data)
                    else:
                        row_data.append(None)
                
                rows_data.append({
                    'row': row_idx,
                    'height': worksheet.row_dimensions[row_idx].height if row_idx in worksheet.row_dimensions else None,
                    'cells': row_data
                })
            
            sheet_data['rows'] = rows_data
            sheet_data['row_count'] = len(rows_data)
            sheet_data['column_count'] = max_col
            
            # Process tables
            for table in worksheet.tables.values():
                try:
                    table_data = self._process_table(table, worksheet)
                    sheet_data['tables'].append(table_data)
                except Exception as e:
                    logger.warning(f"Error processing table {table.name}: {e}")
            
            return sheet_data
            
        except Exception as e:
            logger.error(f"Error processing worksheet {worksheet.title}: {e}")
            raise
    
    def _process_cell(self, cell: Cell) -> Dict[str, Any]:
        """Process a single cell and extract its content and formatting."""
        cell_data = {
            'address': cell.coordinate,
            'value': cell.value,
            'data_type': type(cell.value).__name__,
            'row': cell.row,
            'column': cell.column,
            'column_letter': cell.column_letter,
            'is_date': isinstance(cell.value, (datetime.datetime, datetime.date, datetime.time)),
        }
        
        if self.include_formulas and cell.data_type == 'f':
            cell_data['formula'] = cell.value
            try:
                cell_data['calculated_value'] = cell.value if cell.data_type == 'n' else str(cell.value)
            except:
                cell_data['calculated_value'] = str(cell.value)
        
        if self.include_formatting:
            cell_data['formatting'] = self._extract_cell_formatting(cell)
        
        return cell_data
    
    def _extract_cell_formatting(self, cell: Cell) -> Dict[str, Any]:
        """Extract formatting information from a cell."""
        formatting = {}
        
        # Font
        if cell.font:
            font: Font = cell.font
            formatting['font'] = {
                'name': font.name,
                'size': font.size,
                'bold': font.bold,
                'italic': font.italic,
                'underline': font.underline,
                'strikethrough': font.strike,
                'color': font.color.rgb if font.color and hasattr(font.color, 'rgb') else str(font.color)
            }
        
        # Fill
        if cell.fill:
            fill: PatternFill = cell.fill
            formatting['fill'] = {
                'patternType': fill.patternType,
                'fgColor': fill.fgColor.rgb if fill.fgColor and hasattr(fill.fgColor, 'rgb') else str(fill.fgColor),
                'bgColor': fill.bgColor.rgb if fill.bgColor and hasattr(fill.bgColor, 'rgb') else str(fill.bgColor)
            }
        
        # Border
        if cell.border:
            border: Border = cell.border
            formatting['border'] = {
                'left': str(border.left),
                'right': str(border.right),
                'top': str(border.top),
                'bottom': str(border.bottom),
                'diagonal': str(border.diagonal)
            }
        
        # Alignment
        if cell.alignment:
            align: Alignment = cell.alignment
            formatting['alignment'] = {
                'horizontal': align.horizontal,
                'vertical': align.vertical,
                'textRotation': align.textRotation,
                'wrapText': align.wrapText,
                'shrinkToFit': align.shrinkToFit,
                'indent': align.indent
            }
        
        # Protection
        if cell.protection:
            prot: Protection = cell.protection
            formatting['protection'] = {
                'locked': prot.locked,
                'hidden': prot.hidden
            }
        
        # Number format
        if cell.number_format:
            formatting['number_format'] = cell.number_format
        
        return formatting
    
    def _process_table(self, table, worksheet) -> Dict[str, Any]:
        """Process an Excel table."""
        table_data = {
            'name': table.name,
            'display_name': table.displayName,
            'ref': table.ref,
            'header_row_count': table.headerRowCount,
            'totals_row_count': table.tableStyleInfo.showLastColumn if table.tableStyleInfo else 0,
            'style': table.tableStyleInfo.name if table.tableStyleInfo else None,
            'columns': [],
            'data': []
        }
        
        # Process columns
        for col in table.tableColumns:
            table_data['columns'].append({
                'name': col.name,
                'id': col.id,
                'totals_row_label': col.totalsRowLabel,
                'totals_row_function': col.totalsRowFunction
            })
        
        # Process data
        for row in worksheet[table.ref]:
            row_data = []
            for cell in row:
                row_data.append(self._process_cell(cell))
            table_data['data'].append(row_data)
        
        return table_data
    
    def _extract_defined_names(self) -> List[Dict[str, Any]]:
        """Extract defined names from the workbook."""
        defined_names = []
        
        if not self.workbook.defined_names:
            return defined_names
        
        for name, def_name in self.workbook.defined_names.items():
            try:
                defined_names.append({
                    'name': name,
                    'comment': def_name.comment,
                    'local_sheet_id': def_name.localSheetId,
                    'value': def_name.value,
                    'destinations': [str(dest) for dest in def_name.destinations]
                })
            except Exception as e:
                logger.warning(f"Error processing defined name '{name}': {e}")
                continue
        
        return defined_names
    
    def get_metadata(self) -> Dict[str, Any]:
        """Extract metadata from the Excel workbook."""
        metadata = super().get_metadata()
        
        if not self.workbook:
            return metadata
        
        try:
            # Add Excel-specific metadata
            metadata.update({
                'creator': self.workbook.properties.creator,
                'title': self.workbook.properties.title,
                'description': self.workbook.properties.description,
                'subject': self.workbook.properties.subject,
                'keywords': self.workbook.properties.keywords,
                'category': self.workbook.properties.category,
                'created': self.workbook.properties.created.isoformat() if self.workbook.properties.created else None,
                'modified': self.workbook.properties.modified.isoformat() if self.workbook.properties.modified else None,
                'last_modified_by': self.workbook.properties.lastModifiedBy,
                'revision': self.workbook.properties.revision,
                'version': self.workbook.properties.version,
                'sheet_names': self.workbook.sheetnames,
                'active_sheet': self.workbook.active.title if self.workbook.active else None,
                'calculation': {
                    'force_full_calc': self.workbook.force_full_calc,
                    'full_calc_on_load': self.workbook.full_calc_on_load,
                    'calc_mode': self.workbook.calculation.calcMode if hasattr(self.workbook, 'calculation') else None,
                    'calc_count': self.workbook.calculation.calcCount if hasattr(self.workbook, 'calculation') else None,
                },
                'properties': {
                    'code_name': self.workbook.code_name,
                    'security': {
                        'lock_structure': self.workbook.security.lockStructure,
                        'lock_windows': self.workbook.security.lockWindows,
                    } if hasattr(self.workbook, 'security') else {}
                }
            })
        except Exception as e:
            logger.warning(f"Error extracting Excel metadata: {e}")
        
        return metadata
